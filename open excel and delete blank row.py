import openpyxl
from copy import deepcopy
# 通过输入excel 地址和名字方式打开工作表格
path = input("please input the excel address:")
filename = input("please input the excel name:")
wb = openpyxl.load_workbook(path+'\\'+filename)
wsname = input("please input the worksheet name:")
ws = wb[wsname]
# 将表格中单元格取消合并单元格
# 获取所有的已合并的单元格信息
    # worksheet.merged_cells获取已经合并单元格的信息；再使用worksheet.unmerge_cells()拆分单元格；
m_list = ws.merged_cells
print(m_list)
cr = []
for m_area in m_list:
        # 合并单元格的起始行坐标、终止行坐标。。。。，
    r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
        # 纵向合并单元格的位置信息提取出
    if r2 - r1 > 0:
        cr.append((r1, r2, c1, c2))
    # 这里注意需要把合并单元格的信息提取出再拆分
merge_cr = deepcopy(cr)
for r in cr:
    ws.unmerge_cells(start_row=r[0], end_row=r[1], start_column=r[2], end_column=r[3])
list1 =["Key"]
rowidx=1
colidx=1
rownum=1
# 找到Key所在列
for i in range(1,5):
    for j in range(1,10):
        if ws.cell(i,j).value in list1:
            colidx = j
            rowidx = i
        else:
            print(ws.cell(i,j).value)
#将Key 列的前面行数以及空值所在整行删除
while  rownum < rowidx:
    ws.delete_rows(rownum)
#由于删除行之后，rownum的值会改变，因此这里不要用 rowidx +=1
    rowidx =rowidx-1
while rowidx < ws.max_row:
    if ws.cell(rowidx,colidx).value == None:
        ws.delete_rows(rowidx)
    else:
        rowidx +=1
oldpath = input("if you want to save the excel in old path or not(type Y or N):")
if oldpath == "Y":
    savepath = path+"\\"+filename
else:
    savepath=input("please input the save excel path and name:")
wb.save(savepath)

